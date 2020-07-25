import requests, os, openpyxl, time
from openpyxl.styles import Font
from openpyxl import load_workbook

'''
Python Application to search hashes against VirusTotal
'''

# Headers for the GET Request towards Virustotal
headers = {
		  "Accept-Encoding": "gzip, deflate",
		  "User-Agent" : "gzip,  My Python requests library example client or username"
		  }

# Filename for your excel sheet
filename = "hashresults.xlsx"

# Create the initial excel sheet if it doesn't exist
if not os.path.exists(filename):
	book = openpyxl.Workbook()
	sheet = book['Sheet']

	xlsheaders = ["Hash", "Result"]
	bold_font = Font(color='00FF0000', bold=True)

	for i in range(1, 3):
		sheet.cell(row=1, column=i).value = xlsheaders[i-1]

	for cell in sheet["1:1"]:
		cell.font = bold_font

	book.save(filename)
	
# Open the list of hashes and reads into an array
# Ensure each hash is on each line
f = open('hashes.txt', 'r')
hashList = f.readlines()
f.close()

# Initializing Variables
xlsxRow = 2
count = 0 

# Loop through array of hashes
for hash in hashList:
	try:
		print(hash)
		# Parameters to send to the Virustotal, change the apikey & resource 
		# apikey is the API Key that is tagged to your account, please do not use mine thx
		# resource is the MD5/SHA1/SHA256 Hash
		params = {'apikey': '26196e909342bce7d74989136709a777f37a3eb647c2c2e53b74e26f7f752f44', 'resource': hash}
		
		# Send request to virustotal
		# params are the parameters to be sent, should contain api key and hash
		# headers are the headers to be sent with the GET request, can be changed accordingly
		response = requests.get('https://www.virustotal.com/vtapi/v2/file/report', params=params, headers=headers)
		  
		# json response from virustotal
		json_response = response.json()
		
		# load the excel workbook for editing
		wb = load_workbook(filename)
		sheet = wb.worksheets[0]
		
		# Add the hash to the first column and 2nd row (the first row will start with 1 which is the 'Header', so start with number 2)
		sheet.cell(row=xlsxRow, column=1).value = hash
		
		# if the hash exists in virustotal
		if "scans" in json_response:
			#if symantec has a detection for the hash
			if json_response['scans']['Symantec']['result']:
				sheet.cell(row=xlsxRow, column=2).value = json_response['scans']['Symantec']['result']
			else:
				# if symantec doesn't detect the hash, print an error
				sheet.cell(row=xlsxRow, column=2).value = 'Not detected by Symantec'
		# if the hash is not in virustotal
		else:
			sheet.cell(row=xlsxRow, column=2).value = 'Hash not in VT'
		
		# increase the row for the next iteration
		xlsxRow += 1
		
		# increase the count for each iteration
		count += 1
		# save the file
		wb.save(filename)
		
		# for every 4 hashes, wait for 65 seconds to reset the limit
		if count % 4 == 0:
			time.sleep(65)
			
	# if any errors print the exception and wait for 60 seconds incase 	it is the limiter from virustotal
	except Exception as e:
		print(e)
		time.sleep(60)